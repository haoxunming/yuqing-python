# -*- coding:UTF-8 -*-
#
# Excel workbook helper
#
# Charles 吴波, 2022-07-17: taken from common.py to reduce import time
#
import sys
sys.path.append('../')  # allow 'shared' to be imported below
import shared.common as c

from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

class ExcelWorkbook():
    def __init__(self) -> None:
        self.xls = Workbook()    # work book
        self.ws = self.xls.active
        self.rn_dict = {self.ws.title: 0}
        self.clrs = [[c.RED, '00FF0000'], [c.YELLOW, '00FFFF00'], [c.PURPLE, '00FF00FF'], [c.GREY, '00C0C0C0']]

    def convert_clr(self, clr):
        return None if clr is None else ([v[1] for v in self.clrs if v[0] == clr][0] if type(clr) is int else clr)
        
    def create_worksheet(self, name=' '):
        self.ws = self.xls.create_sheet(name)
        self.rn_dict[name] = 0
        return self.ws

    def delete_worksheet(self):
        self.xls.remove(self.ws)
        self.ws = None
        
    def get_max_rows(self, ws = None):
        if not ws: ws = self.ws
        return ws.max_row if ws else 0

    def get_row_count(self, ws = None):
        if not ws: ws = self.ws
        return self.rn_dict[ws.title]

    def get_font(self, clr = None, bold = None):
        return Font(color=self.convert_clr(clr), bold=bold)

    def get_fill(self, clr = None, fill_type = None):
        clr = self.convert_clr(clr)
        return PatternFill(start_color=clr, end_color=clr, fill_type=fill_type)

    def set_title(self, title, ws = None):
        if not ws: ws = self.ws
        rn = self.rn_dict[ws.title]
        del self.rn_dict[ws.title]
        ws.title = title
        self.rn_dict[title] = rn
        return ws

    def set_columns(self, cols: list, ws = None) -> None:
        col = 1
        if not ws: ws = self.ws
        self.rn_dict[ws.title] += 1
        for name, width in cols:
            ws.cell(row=self.rn_dict[ws.title], column=col).value = name
            if width:
                ws.column_dimensions[get_column_letter(col)].width = width
            col += 1

    def set_cell(self, col_ix: int, value: object, font: Font = None, fill: PatternFill = None, halign: list = None, valign: list = None, ws = None) -> None:
        if not ws: ws = self.ws
        if ws:
            cell = ws.cell(row=self.rn_dict[ws.title], column=col_ix)
            cell.value = ILLEGAL_CHARACTERS_RE.sub('', value) if type(value) is str else value
            if font:
                cell.font = font

            if fill:
                cell.fill = fill

            if halign or valign:
                col_ix -= 1
                cell.alignment = Alignment(horizontal=halign[col_ix] if halign and col_ix < len(halign) else None, vertical=valign[col_ix] if valign and col_ix < len(valign) else None)


    def add_row(self, values: list = None, start_col=1, font=None, fill=None, halign=None, valign=None, ws = None) -> None:
        if not ws: ws = self.ws
        self.rn_dict[ws.title] += 1
        if values:
            for ix, text in enumerate(values):
                self.set_cell(start_col + ix, text, font, fill, halign, valign, ws)  
        return self.rn_dict[ws.title]      

    def save(self, file_name:str, open=False) -> bool:
        try:
            self.xls.save(file_name)
        except Exception as e:
            a = input(f'错误: 无法保存 {file_name}: {repr(e)}. 按 ENTER 键重新尝试, 其他键退出')
            if a:
                return False
            try:
                self.xls.save(file_name)
            except Exception as e:
                a = input(f'错误: 无法再次保存 {file_name}: {repr(e)}. 按 ENTER 键退出')
                return False

        if open:
            c.startfile(file_name)
        return True
